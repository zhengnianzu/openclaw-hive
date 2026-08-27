# -*- coding: utf-8 -*-
"""批跑作业「输出」页端点：只读写 platform.db（task_records + task_traj_records），不碰 obsutil。

数据源（README_outputview.md §二-0）：复用 platform.db，不另起 hive_output.db。
  - 浅层 = task_records（唯一事实表）：漏斗分布按 traj_level SQL 聚合；会话表格直接查行。
  - 深层 = task_traj_records：行由 offline/output_worker.py 回填分级；用户点开详情时
    POST 置 status='pending'，worker 消费后回填 5 个路径列（status → done/failed）。

全部端点同步实现（async_execute 跑在专用线程池），绝不阻塞事件循环，也不触发 OBS 下载。
"""
from __future__ import annotations

import json
import os
import re
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

from ..core.config import settings
from ..core.database import get_connection, async_execute, async_query, async_query_one
from ..core.security import get_current_user, require_operator

router = APIRouter(prefix="/api/instances", tags=["batch-output"])

# proxy_name.json：IP → 代理名映射（platform/settings/proxy_name.json），不存在则映射为空
_PROXY_NAME_FILE = os.path.join(settings.SETTINGS_DIR, "proxy_name.json")
_IP_RE = re.compile(r"https?://([0-9.]+)")

# 面板角色：主Agent 取 create_params 顶层，用户模拟/Evaluator 取 agents[] 对应项
_MAIN_AGENT_ROLE = "主Agent"
_USER_SIM_ROLE = "用户模拟"
_EVALUATOR_ROLE = "Evaluator"

# 角色 → create_params 里的字段定位（供对话导出：取完整 key + model）
_MAIN_AGENT = "main_agent"
_USER_SIM = "user_simulator"
_EVALUATOR = "evaluator"
_ROLE_TO_KEY = {
    _MAIN_AGENT_ROLE: (_MAIN_AGENT, "model_api_key", "model_base_url", "model_id"),
    _USER_SIM_ROLE: (_USER_SIM, "api_key", "base_url", "model"),
    _EVALUATOR_ROLE: (_EVALUATOR, "api_key", "base_url", "model"),
}

# 导出方式（外部服务允许的 mode）+ 默认 mode
_EXPORT_MODES = ["export", "reformat", "eval", "reconstruct", "full_reformat"]
_DEFAULT_MODE = "export"

# 导出任务状态：外部 queued/running/success/failed + 本地未导出
_EXPORT_STATUS_ACTIVE = {"queued", "running"}
_EXPORT_STATUS_FINAL = {"success", "failed"}

# 漏斗档位（x 轴顺序，与 README 一致）
_TRAJ_LEVELS = ["L0", "L1", "L1.5", "L2", "L3"]

# 会话表格允许的排序列（防 SQL 注入）——task_records 的实际列
_SORTABLE = {
    "task_idx", "config_name", "traj_level", "status", "eval_score", "eval_completion",
}


def _get_instance(instance_id: str) -> dict:
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM task_instances WHERE id=?", (instance_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="实例不存在")
    return dict(row)


# ============ 浅层 ============

@router.post("/{instance_id}/shallow")
async def trigger_shallow(instance_id: str, user: dict = Depends(require_operator)):
    """请求浅层分级：登记该实例到 shallow_requests 表，worker 下一轮消费。

    点击驱动：worker 不再全量扫历史 finished 缺口实例（动辄数十万行追不完），
    只消费 running/preparing + 本表登记的 finished/completed 实例。幂等：重复登记只更新
    status/created_by，不动已有的处理进度。处理完成后 worker 删登记出队。

    output_status 门控：任务完成（completed）且 output_status='done' 时说明浅层已完成、
    结果可直接用 → 不再重复登记，直接返回；output_status IS NULL 才允许提交新浅层。
    """
    inst = _get_instance(instance_id)
    if inst.get("output_status") == "done":
        return {"instance_id": instance_id, "status": "done",
                "hint": "该实例浅层已完成(output_status=done)，直接使用，无需重复处理"}
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO shallow_requests (instance_id, created_by, status, created_at) "
            "VALUES (?, ?, 'queued', CURRENT_TIMESTAMP) "
            "ON CONFLICT(instance_id) DO UPDATE SET "
            "status='queued', created_by=excluded.created_by, created_at=CURRENT_TIMESTAMP",
            (instance_id, user.get("username")),
        )
    return {"instance_id": instance_id, "status": "queued",
            "hint": "已登记浅层请求，worker 下一轮消费，GET /shallow 观察进度"}


@router.get("/{instance_id}/shallow")
async def get_shallow(instance_id: str, user: dict = Depends(get_current_user)):
    """浅层统计 + 进度（不再返回会话行，表格数据走 /shallow/tasks 分页接口）。

    返回 {summary, progress}：summary 为 L0-L3 计数（traj_level 分组），progress 为
    浅层处理进度。会话行已由 /shallow/tasks（分页+过滤+排序）承载，避免全量行
    每 10s 传输约 1MB 造成页面卡顿。
    """
    inst = _get_instance(instance_id)

    with get_connection() as conn:
        rows = conn.execute(
            "SELECT tr.traj_level, ttr.task_done AS ttr_task_done "
            "FROM task_records tr "
            "LEFT JOIN task_traj_records ttr "
            "  ON ttr.instance_id = tr.instance_id AND ttr.config_name = tr.config_name "
            "WHERE tr.instance_id=? ORDER BY tr.task_idx",
            (instance_id,),
        ).fetchall()
        rows = [dict(r) for r in rows]
        graded = [r for r in rows if r["traj_level"] in _TRAJ_LEVELS]

    summary = {lv: 0 for lv in _TRAJ_LEVELS}
    for r in graded:
        summary[r["traj_level"]] += 1
    summary["graded"] = len(graded)
    summary["total"] = len(rows)
    summary["task_done"] = sum(1 for r in rows if r.get("ttr_task_done"))
    # 未评估 = traj_level 为 NULL（未被浅层分析覆盖/队列未消费）
    summary["unevaluated"] = sum(1 for r in rows if r["traj_level"] is None)

    # 浅层进度：已分级 + failed 占位 = 处理完；残留 NULL = 未处理。
    # finished 实例已完成分级（历史堆积）时 both = total，进度条隐藏。
    # unaudited = 仍为 NULL 的行数（浅层分析未覆盖/队列未消费）。
    unaudited = len(rows) - len(graded) - sum(
        1 for r in rows if r["traj_level"] == "failed")
    queued = False
    status = inst.get("status")
    if status in ("finished", "completed", "stopped"):
        # finished/completed/stopped 实例是否被登记（worker 待处理或正在处理）
        with get_connection() as conn:
            queued = conn.execute(
                "SELECT 1 FROM shallow_requests WHERE instance_id=? LIMIT 1",
                (instance_id,),
            ).fetchone() is not None

    return {
        "instance_id": instance_id,
        "instance_status": status,
        "output_status": inst.get("output_status"),
        "total_tasks": inst.get("total_tasks"),
        "summary": summary,
        # 浅层进度：processed = graded + failed（已定级）；undone = total - processed；
        # queued = ended 实例且已登记（worker 待处理）；全部行处理后 undone=0。
        "progress": {
            "approved": len(graded),
            "failed": sum(1 for r in rows if r["traj_level"] == "failed"),
            "total": len(rows),
            "undone": unaudited,
            "queued": queued,
        },
    }


@router.get("/{instance_id}/shallow/tasks")
async def shallow_tasks(
    instance_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    tag: str | None = Query(None, description="L0/L1/L1.5/L2/L3/done/fail/unevaluated"),
    keyword: str | None = Query(None, description="按 config_name 模糊搜索"),
    sort_by: str = Query("task_idx"),
    sort_dir: str = Query("asc"),
    user: dict = Depends(get_current_user),
):
    """会话表格数据（分页 + 标签过滤 + 关键词 + 排序）。"""
    inst = _get_instance(instance_id)

    where = ["tr.instance_id = ?"]
    params: list = [instance_id]
    if tag:
        if tag == "done":
            where.append("tr.traj_level IS NOT NULL")
        elif tag == "fail":
            where.append("tr.traj_level = 'failed'")
        elif tag == "unevaluated":
            where.append("tr.traj_level IS NULL")
        elif tag in _TRAJ_LEVELS:
            where.append("tr.traj_level = ?")
            params.append(tag)
    if keyword:
        kw = keyword.strip()
        if kw:
            where.append("tr.config_name LIKE ?")
            params.append(f"%{kw}%")
    where_sql = " AND ".join(where)

    # 排序列限定到白名单；JOIN 后两表均有 task_idx 列，裸列名会歧义，统一加 tr. 前缀。
    # （_SORTABLE 全部是 task_records 的列。）
    col = sort_by if sort_by in _SORTABLE else "task_idx"
    direction = "DESC" if str(sort_dir).lower() == "desc" else "ASC"

    from_sql = ("FROM task_records tr "
                "LEFT JOIN task_traj_records ttr "
                "  ON ttr.instance_id = tr.instance_id AND ttr.config_name = tr.config_name")
    total_row = await async_query_one(
        f"SELECT COUNT(*) AS n {from_sql} WHERE {where_sql}", tuple(params)
    )
    total = total_row["n"] if total_row else 0
    offset = (page - 1) * page_size
    tasks = await async_query(
        f"SELECT tr.task_idx, tr.config_name, tr.traj_level, tr.status, tr.error_code, "
        f"tr.error_category, tr.eval_score, tr.eval_completion, tr.gate, "
        f"ttr.task_done, ttr.has_eval "
        f"{from_sql} WHERE {where_sql} ORDER BY tr.{col} {direction} LIMIT ? OFFSET ?",
        tuple(params) + (page_size, offset),
    )
    return {"total": total, "page": page, "page_size": page_size, "tasks": tasks}


# ============ 深层 ============

# 漏斗等级（批量入队的合法 traj_level；failed / NULL 不入队）
_ENQUEUE_LEVELS = _TRAJ_LEVELS

# ⚠ 静态批量端点（enqueue_all/queue）必须定义在参数路由 {traj_name} 之前，
#   否则 FastAPI 按注册序匹配，POST /deep/enqueue_all 会被 trigger_deep 的
#   {traj_name} 捕获（traj_name="enqueue_all"）→ 404。


@router.post("/{instance_id}/deep/enqueue_all")
async def deep_enqueue_all(instance_id: str, user: dict = Depends(require_operator)):
    """把该实例所有已分级（L0-L3）且未下载的行批量置 pending 入队。

    与单会话 trigger_deep 同款幂等保护：
      - 已 downloading 的不打断（跳过）
      - 已有 assistant_traj_path（此前下载成功）的跳过，不重复拉 OBS
      - failed / level IS NULL 的行不参与（无轨迹或不可分级）
    返回 {queued, skipped_done, skipped_downloading, skipped_ungraded, total}。
    """
    inst = _get_instance(instance_id)
    level_ph = ",".join("?" for _ in _ENQUEUE_LEVELS)
    with get_connection() as conn:
        # 已 done 且已下载过（有路径）——批量入队时跳过，不重复下载
        skipped_done = conn.execute(
            "SELECT COUNT(*) FROM task_traj_records "
            "WHERE instance_id=? AND assistant_traj_path IS NOT NULL AND status='done'",
            (instance_id,),
        ).fetchone()[0]
        # 正在下载的——不打断
        skipped_downloading = conn.execute(
            "SELECT COUNT(*) FROM task_traj_records "
            "WHERE instance_id=? AND status='downloading'",
            (instance_id,),
        ).fetchone()[0]
        # 入队：L0-L3 且未下载（无路径）且未 downloading 的行 → pending
        cur = conn.execute(
            f"UPDATE task_traj_records SET status='pending', updated_at=CURRENT_TIMESTAMP "
            f"WHERE instance_id=? AND level IN ({level_ph}) "
            f"AND assistant_traj_path IS NULL AND status NOT IN ('downloading')",
            (instance_id, *_ENQUEUE_LEVELS),
        )
        conn.commit()
    queued = cur.rowcount or 0
    with get_connection() as conn:
        skipped_ungraded = conn.execute(
            "SELECT COUNT(*) FROM task_traj_records "
            "WHERE instance_id=? AND (level IS NULL OR level='failed') "
            "AND assistant_traj_path IS NULL",
            (instance_id,),
        ).fetchone()[0]
    return {
        "instance_id": instance_id,
        "queued": queued,
        "skipped_done": skipped_done,
        "skipped_downloading": skipped_downloading,
        "skipped_ungraded": skipped_ungraded,
        "total": queued + skipped_done + skipped_downloading + skipped_ungraded,
        "status": "queued" if queued else "noop",
        "hint": "worker 常驻消费 pending 行，GET /deep/{traj}/status 或 /deep/queue 观察进度",
    }


@router.get("/{instance_id}/deep/queue")
async def deep_queue(instance_id: str, page: int = 1, page_size: int = 50,
                     user: dict = Depends(get_current_user)):
    """该实例深层下载队列状态（前端轮询）。

    返回 {summary, rows, total}：
      summary = {pending, downloading, done, failed, total}（该实例 task_traj_records 计数）
      rows    = 按「进行中优先」排序的分页行（pending/downloading 在前，再 failed/done），
                每行含 traj_name/level/status/updated_at/error/assistant_traj_path。
      total   = 满足排序的全部行数（用于前端分页 total）。
    """
    inst = _get_instance(instance_id)
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    off = (page - 1) * page_size
    with get_connection() as conn:
        # summary 独立 COUNT，避免为拿计数而返回全量行（大实例全量行很重）
        counted = conn.execute(
            "SELECT status, COUNT(*) AS n FROM task_traj_records "
            "WHERE instance_id=? GROUP BY status", (instance_id,),
        ).fetchall()
        total = 0
        summary = {"pending": 0, "downloading": 0, "done": 0, "failed": 0}
        for c in counted:
            n = c["n"]
            total += n
            if c["status"] in summary:
                summary[c["status"]] = n
        summary["total"] = total

        rows = conn.execute(
            "SELECT traj_name, level, status, error, assistant_traj_path, updated_at "
            "FROM task_traj_records WHERE instance_id=? "
            "ORDER BY CASE status "
            "           WHEN 'pending' THEN 0 WHEN 'downloading' THEN 1 "
            "           WHEN 'failed' THEN 2 ELSE 3 END, updated_at DESC "
            "LIMIT ? OFFSET ?",
            (instance_id, page_size, off),
        ).fetchall()
        rows = [dict(r) for r in rows]

    return {"instance_id": instance_id, "summary": summary, "rows": rows, "total": total}


@router.post("/{instance_id}/deep/{traj_name}")
async def trigger_deep(instance_id: str, traj_name: str,
                       user: dict = Depends(require_operator)):
    """请求深层加载：该 task 的 task_traj_records 行置 pending，worker 下载并回填路径列。"""
    inst = _get_instance(instance_id)
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id, status, assistant_traj_path FROM task_traj_records "
            "WHERE instance_id=? AND traj_name=?",
            (instance_id, traj_name),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404,
                                detail=f"无轨迹记录: {traj_name}（浅层分级可能尚未完成）")
        # 路径列已齐全（此前下载成功过）→ 恢复 done，不重复下载
        if row["assistant_traj_path"]:
            conn.execute(
                "UPDATE task_traj_records SET status='done', updated_at=CURRENT_TIMESTAMP "
                "WHERE id=?",
                (row["id"],),
            )
            return {"instance_id": instance_id, "traj_name": traj_name,
                    "status": "done"}
        # 已在 downloading/进行中则幂等跳过（不打断正在下载的）
        cur = conn.execute(
            "UPDATE task_traj_records SET status='pending', updated_at=CURRENT_TIMESTAMP "
            "WHERE id=? AND status NOT IN ('downloading')",
            (row["id"],),
        )
        already = cur.rowcount == 0
    return {"instance_id": instance_id, "traj_name": traj_name,
            "status": "queued" if not already else "in_progress"}


@router.get("/{instance_id}/deep/{traj_name}/status")
async def deep_status(instance_id: str, traj_name: str,
                      user: dict = Depends(get_current_user)):
    """深层加载状态：前端 3s 轮询。done 时附带 5 个缓存路径。"""
    inst = _get_instance(instance_id)
    with get_connection() as conn:
        row = conn.execute(
            "SELECT traj_name, status, error, assistant_traj_path, evaluator_traj_path, "
            "task_log_path, gateway_log_path, eval_log_path "
            "FROM task_traj_records WHERE instance_id=? AND traj_name=?",
            (instance_id, traj_name),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=f"无轨迹记录: {traj_name}")
    return dict(row)


@router.get("/{instance_id}/deep/{traj_name}/detail")
async def deep_detail(instance_id: str, traj_name: str,
                      user: dict = Depends(get_current_user)):
    """读取本地缓存的详情（assistant/evaluator 轨迹块 + 主 log/gateway/eval log 尾部）。

    需先 POST /deep/{traj_name} 触发下载、worker 状态 done 后调用；本地缓存无数据返回 409，
    前端据 status 轮询后重试。只读 output_cache/<batch>/<traj_name>/，不触发 OBS 下载。
    """
    inst = _get_instance(instance_id)
    with get_connection() as conn:
        tr = conn.execute(
            "SELECT * FROM task_traj_records WHERE instance_id=? AND traj_name=?",
            (instance_id, traj_name),
        ).fetchone()
    if not tr:
        raise HTTPException(status_code=404, detail=f"无轨迹记录: {traj_name}")
    tr = dict(tr)

    cache_root = os.path.abspath(settings.OUTPUT_CACHE)
    # 任务目录判定优先级：
    #   1. assistant_traj_path（worker 深层下载后回填的缓存绝对路径，obsutil cp 保留
    #      OBS leaf 目录 → "<batch>/<leaf>/<leaf>/projects/..."，取 traj_name 最后出现处前缀）
    #   2. trajectory_rel（浅层写入的 OBS 路径，<batch...>/<leaf>/...，取 leaf 前缀）
    task_dir = None
    for rel in (tr.get("assistant_traj_path"), tr.get("trajectory_rel")):
        if not rel:
            continue
        rparts = (rel.replace(cache_root, "", 1).lstrip("/") if rel.startswith(cache_root)
                  else rel).split("/")
        # traj_name 在 rel 中第一次出现处：任务目录 = cache_root + 此前缀。
        # 注：obsutil cp -r 会保留 OBS leaf 目录 → 深层下载后本地是 <batch>/<leaf>/<leaf>/… 双层嵌套，
        #     assistant_traj_path 指向子层；但日志（run.log/gateway.log/evaluator_use.log）落在父层 dest。
        #     故取「第一次出现」（=父层任务根）而非最后一次（=子层），否则读不到父层日志。
        #     浅层 trajectory_rel 为 <batch>/<leaf>/agents/… 时第一次出现即任务根，同样正确。
        idx = rparts.index(tr["traj_name"]) if tr["traj_name"] in rparts else -1
        if idx >= 0:
            candidate = os.path.join(cache_root, *rparts[:idx + 1])
            if os.path.isdir(candidate):
                task_dir = candidate
                break
    if task_dir is None:
        # 退化：直接以 traj_name 为目录名
        candidate = os.path.join(cache_root, tr["traj_name"])
        if os.path.isdir(candidate):
            task_dir = candidate
    if task_dir is None:
        raise HTTPException(status_code=409,
                            detail="本地缓存不存在，请先 POST /deep/{traj_name} 触发下载")

    try:
        from src import offline_analysis as oa
        detail = oa.load_task_detail(task_dir, traj_name,
                                     known_harness=tr.get("harness"))
        # 本地是否有真实 assistant 轨迹文件（多候选取最大者；任务确无回复时 OBS 无轨迹，为 None）
        has_traj_file = bool(oa.find_primary_assistant_trajectory(task_dir))
        if has_traj_file:
            # 轨迹文件在但 stats 不可解析，且无 tsr 兜底 → 真异常
            if not detail.get("assistant_stats") and not oa.read_tsr_stats(task_dir):
                raise HTTPException(status_code=500,
                                    detail=f"本地详情解析失败: 缓存目录 {task_dir} 无轨迹也无 stats")
        else:
            # 本地无轨迹文件。区分两种情形：
            #   deep 已下载过（worker 会把可见产物回填到路径列；shallow 只拉 tsr 不碰主 log，
            #   故仅浅层快照时五项路径列全 NULL 且本地无主 log）→ 任务确无 assistant 轨迹
            #   （工具全失败/无助手回复）→ 不 409，tsr 兜底展示，前端渲染"无 assistant 轨迹"。
            #   未触发 deep、无任何深层产物 → 本地仅 tsr 浅层快照 → 409 触发下载。
            # 注：task_done 不作判据——它标"任务是否完成"，而 done 行里大量 task_done=0
            #   的浅层产物同样无文件、轨迹仍在 OBS，仍需触发下载。
            deep_ran = any(tr.get(k) for k in (
                "assistant_traj_path", "evaluator_traj_path",
                "task_log_path", "gateway_log_path", "eval_log_path"))
            if deep_ran or oa.find_primary_log(task_dir, traj_name):
                if not detail.get("assistant_stats") and not oa.read_tsr_stats(task_dir):
                    raise HTTPException(status_code=500,
                                        detail=f"本地详情解析失败: 缓存目录 {task_dir} 无轨迹也无 stats")
            else:
                raise HTTPException(status_code=409,
                                    detail="本地轨迹未下载，请先加载详情触发下载")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"本地详情解析失败: {str(e)[:200]}")

    # 路径列再确认（trajectory_rel 可能因缓存目录结构偏差需调整）
    if not tr.get("assistant_traj_path"):
        for t in detail.get("trajectories") or []:
            if t.get("role") == "assistant" and not tr.get("assistant_traj_path"):
                tr["assistant_traj_path"] = t["path"]
            elif t.get("role") == "evaluator" and not tr.get("evaluator_traj_path"):
                tr["evaluator_traj_path"] = t["path"]

    # 深层下载状态：以 DB 5 个路径列是否回填为准（worker 下载成功后落列）。
    # 全 NULL = 深层从未成功回填（即便本地残留旧版缓存的轨迹文件），判 not_downloaded，
    # 前端据 deep_status 触发重新下载补齐（网关/eval 日志随重构后的下载一并落盘并回填）。
    deep_paths = [tr.get(k) for k in (
        "assistant_traj_path", "evaluator_traj_path",
        "task_log_path", "gateway_log_path", "eval_log_path")]
    deep_status = ("downloaded" if any(deep_paths) else "not_downloaded")

    return {
        "traj_name": traj_name,
        "harness": tr.get("harness") or detail.get("harness"),
        "deep_status": deep_status,
        "deep_error": tr.get("error"),
        "assistant_stats": detail.get("assistant_stats"),
        "assistant_trajectory": detail.get("assistant_trajectory"),
        "evaluator_trajectory": detail.get("evaluator_trajectory"),
        "trajectories": detail.get("trajectories"),
        "log": detail.get("log"),
        "gateway": detail.get("gateway"),
        "eval_use_log": detail.get("eval_use_log"),
        "verdict": detail.get("verdict"),
        "paths": {
            "assistant_traj_path": tr.get("assistant_traj_path"),
            "evaluator_traj_path": tr.get("evaluator_traj_path"),
            "task_log_path": tr.get("task_log_path"),
            "gateway_log_path": tr.get("gateway_log_path"),
            "eval_log_path": tr.get("eval_log_path"),
        },
    }


# ============ 模型配置 / 用户模拟配置信息面板 ============
# 数据来源：task_instances.create_params（创建实例时录入的表单快照），不读实例目录文件。
# 主Agent = 顶层 model_base_url/model_api_key/model_id；
# 用户模拟 = agents[] 中 name != 'evaluator' 的项（可多行）；
# Evaluator = agents[] 中 name == 'evaluator' 的项。


def _load_proxy_map() -> dict:
    """读取 proxy_name.json（IP → 代理名）。文件缺失/解析失败/非 dict 时返回空字典（映射为空）。"""
    try:
        with open(_PROXY_NAME_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (OSError, ValueError):
        return {}


def _proxy_name_for(baseurl: str, proxy_map: dict) -> str:
    """从 baseurl 提取 IP 查代理名；无 IP / 未命中 / 空映射 → 返回空串。"""
    if not baseurl:
        return ""
    m = _IP_RE.search(baseurl)
    if not m:
        return ""
    return proxy_map.get(m.group(1), "")


def _key_suffix(api_key) -> str:
    """key 只取后四位；空或非字符串返回空串。"""
    if not api_key or not isinstance(api_key, str):
        return ""
    return api_key[-4:] if len(api_key) >= 4 else api_key


def _build_proxy_rows(create_params) -> list:
    """从 create_params dict 拼出面板 rows（key 已截后四位、代理名已映射）。"""
    if not create_params or not isinstance(create_params, dict):
        return []
    proxy_map = _load_proxy_map()
    rows = []

    # 主Agent：顶层字段
    rows.append({
        "role": _MAIN_AGENT_ROLE,
        "baseurl": create_params.get("model_base_url") or "",
        "key_suffix": _key_suffix(create_params.get("model_api_key")),
        "proxy_name": _proxy_name_for(create_params.get("model_base_url"), proxy_map),
        "model": create_params.get("model_id") or "",
    })

    # 用户模拟 / Evaluator：agents[]（name 决定角色）
    for a in (create_params.get("agents") or []):
        if not isinstance(a, dict):
            continue
        name = a.get("name") or ""
        role = _EVALUATOR_ROLE if name == "evaluator" else _USER_SIM_ROLE
        rows.append({
            "role": role,
            "baseurl": a.get("base_url") or "",
            "key_suffix": _key_suffix(a.get("api_key")),
            "proxy_name": _proxy_name_for(a.get("base_url"), proxy_map),
            "model": a.get("model") or "",
        })

    return rows


# ============ 模型配置面板：对话导出（对接外部 EXPORT_BASE 服务） ============
# 说明：浏览/导出提交/状态查询都经本后端代理，完整 key + model + access-key 只在服务端拼装，
# 前端只接触 rows 里的 key_suffix（后四位）。外部服务契约见 doc/README_export.md / 用户提供文档。

def _agent_for_role(create_params: dict, role: str) -> dict:
    """按角色从 create_params 定位完整连接信息 dict（key/baseurl/model）。

    找不到该角色时返回空 dict（调用方判定 404）。主Agent 取顶层字段；
    用户模拟/Evaluator 取 agents[] 里 name 匹配项（agents[].name 是 user_simulator/evaluator，
    映射到显示角色，与 _build_proxy_rows 的分组逻辑一致）。
    """
    if not create_params or not isinstance(create_params, dict):
        return {}
    if role == _MAIN_AGENT_ROLE:
        return {
            "key": create_params.get("model_api_key") or "",
            "baseurl": create_params.get("model_base_url") or "",
            "model": create_params.get("model_id") or "",
        }
    # 显示角色 → agents[].name：evaluator 命中，其余非主Agent 角色按 user_simulator 处理
    target_name = "evaluator" if role == _EVALUATOR_ROLE else "user_simulator"
    for a in (create_params.get("agents") or []):
        if not isinstance(a, dict):
            continue
        if (a.get("name") or "") == target_name:
            return {
                "key": a.get("api_key") or "",
                "baseurl": a.get("base_url") or "",
                "model": a.get("model") or "",
            }
    return {}


def _export_status_map(instance_id: str) -> dict:
    """读 model_export_tasks，返回 {role: {...}}（未导出的角色不在 map 里）。"""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT role, export_id, export_name, mode, status, session_path, "
            "total_sessions, error_message FROM model_export_tasks WHERE instance_id=?",
            (instance_id,),
        ).fetchall()
    out = {}
    for r in rows:
        out[r["role"]] = dict(r)
    return out


def _upsert_export_task(instance_id: str, role: str, model_key: str, model: str,
                        export_id: int, export_name: str, mode: str, status: str,
                        session_path: str = "", total_sessions: int = None,
                        error_message: str = ""):
    """UPSERT 一条导出任务（UNIQUE(instance_id, role)）。update 时刷新 updated_at。"""
    with get_connection() as conn:
        cur = conn.execute(
            "SELECT id FROM model_export_tasks WHERE instance_id=? AND role=?",
            (instance_id, role),
        ).fetchone()
        if cur:
            conn.execute(
                "UPDATE model_export_tasks SET model_key=?, model=?, export_id=?, export_name=?, "
                "mode=?, status=?, session_path=?, total_sessions=?, error_message=?, "
                "updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (model_key, model, export_id, export_name, mode, status,
                 session_path, total_sessions, error_message, cur["id"]),
            )
        else:
            conn.execute(
                "INSERT INTO model_export_tasks (instance_id, role, model_key, model, export_id, "
                "export_name, mode, status, session_path, total_sessions, error_message) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (instance_id, role, model_key, model, export_id, export_name, mode,
                 status, session_path, total_sessions, error_message),
            )


def _export_url(path: str, params: dict) -> str:
    """拼外部服务完整 URL：base（去尾斜杠）+ path + query（已 URL-encode）。

    保留空值参数（如 model=），外部契约要求 key/model 参数始终出现在 query 里。
    """
    base = (settings.EXPORT_BASE or "").rstrip("/")
    qs = "&".join(f"{k}={quote(str(v))}" for k, v in params.items() if v is not None)
    return f"{base}{path}?{qs}" if qs else f"{base}{path}"


async def _call_export(path: str, params: dict, timeout: float = 30.0) -> dict:
    """调外部导出服务。状态码非 2xx 时抛 HTTPException（透传外部 detail）。

    返回 JSON dict。外部 403/404/400 的三类 detail（Invalid access-key / Not found / Invalid mode）
    原样透传给前端，语义与契约一致。
    """
    if not settings.EXPORT_BASE or not settings.EXPORT_ACCESS_KEY:
        raise HTTPException(status_code=500, detail="导出服务未配置（EXPORT_BASE / EXPORT_ACCESS_KEY）")
    params = {**params, "access-key": settings.EXPORT_ACCESS_KEY}
    url = _export_url(path, params)
    try:
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=False) as client:
            resp = await client.get(url)
    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"导出服务不可达: {e}")
    if resp.status_code >= 400:
        detail = "导出服务返回错误"
        try:
            body = resp.json()
            if isinstance(body, dict) and body.get("detail"):
                detail = body["detail"]
        except ValueError:
            pass
        raise HTTPException(status_code=resp.status_code, detail=detail)
    try:
        return resp.json()
    except ValueError:
        raise HTTPException(status_code=502, detail="导出服务返回非 JSON 响应")


@router.get("/{instance_id}/proxy-config")
async def get_proxy_config(instance_id: str, user: dict = Depends(get_current_user)):
    """模型配置/用户模拟配置信息面板数据。

    读 task_instances.create_params → 主Agent/用户模拟/Evaluator 三组 rows
    （baseurl / key 后四位 / 代理名称 / 模型），key 已截尾，代理名由 settings/proxy_name.json
    做 IP→名称映射（无映射留空）。运行时间取实例级 started_at ~ stopped_at。
    每行并回带对话导出信息（export_status / session_path / total_sessions，来自 model_export_tasks）。
    """
    inst = _get_instance(instance_id)
    rows = []
    try:
        if inst.get("create_params"):
            rows = _build_proxy_rows(json.loads(inst["create_params"]))
    except (ValueError, TypeError):
        rows = []  # create_params 非法时静默降级为空，不 500

    # 合并对话导出状态：role → model_export_tasks 行
    export_map = _export_status_map(instance_id)
    for r in rows:
        ex = export_map.get(r["role"])
        r["export_status"] = (ex or {}).get("status") or "unexported"
        r["session_path"] = (ex or {}).get("session_path") or ""
        r["total_sessions"] = (ex or {}).get("total_sessions")
        r["export_id"] = (ex or {}).get("export_id")
        r["export_name"] = (ex or {}).get("export_name") or ""
        r["export_mode"] = (ex or {}).get("mode") or ""

    return {
        "instance_id": instance_id,
        "name": inst.get("name") or "",
        "status": inst.get("status") or "",
        "started_at": inst.get("started_at") or "",
        "stopped_at": inst.get("stopped_at") or "",
        "rows": rows,
    }


@router.get("/{instance_id}/proxy-config/export")
async def export_proxy_config(instance_id: str, user: dict = Depends(get_current_user)):
    """导出模型配置面板为 .xlsx（openpyxl 生成，FileResponse 下载）。"""
    inst = _get_instance(instance_id)
    rows = []
    try:
        if inst.get("create_params"):
            rows = _build_proxy_rows(json.loads(inst["create_params"]))
    except (ValueError, TypeError):
        rows = []

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "模型配置"
    headers = ["名称", "baseurl", "key 后四位", "代理名称", "启动时间", "结束时间", "模型"]
    ws.append(headers)
    # 表头样式：加粗 + 灰底
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    started = inst.get("started_at") or ""
    stopped = inst.get("stopped_at") or ""
    for r in rows:
        ws.append([
            r["role"],
            r["baseurl"],
            r["key_suffix"],
            r["proxy_name"],
            started,
            stopped,
            r["model"],
        ])

    # 列宽自适应（按内容最大宽 + 余量）
    for col_idx, cell in enumerate(ws[1], start=1):
        max_len = len(str(cell.value or ""))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in row:
                max_len = max(max_len, len(str(c.value or "")))
        ws.column_dimensions[cell.column_letter].width = max_len + 6

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = (inst.get("name") or instance_id).replace("/", "_").replace('"', "")
    filename = f"{safe_name}_proxy_config.xlsx"
    # HTTP 头只能用 latin-1，中文文件名会 UnicodeEncodeError。
    # 用 RFC 5987 的 filename*（UTF-8 百分号编码）承载中文名，
    # 再给一个纯 ASCII 的 filename 兜底（非 ASCII 字符替换为下划线）。
    from urllib.parse import quote
    ascii_fallback = filename.encode("ascii", "replace").decode("ascii").replace("?", "_")
    disposition = (
        f"attachment; filename=\"{ascii_fallback}\"; "
        f"filename*=UTF-8''{quote(filename)}"
    )
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"content-disposition": disposition},
    )


# ============ 模型配置面板：对话导出端点 ============

def _role_cred_from(create_params, role: str, inst: dict):
    """按角色取完整连接信息；create_params 非法或该角色缺失时抛 404。

    返回 (key, model)。供导出提交/浏览/状态查询定位该 row 的凭据。
    """
    cp = None
    try:
        if inst.get("create_params"):
            cp = json.loads(inst["create_params"])
    except (ValueError, TypeError):
        cp = None
    info = _agent_for_role(cp, role)
    if not info or not info.get("key"):
        raise HTTPException(status_code=404, detail=f"角色 {role} 未找到或缺失 key")
    return info["key"], info.get("model") or ""


@router.post("/{instance_id}/proxy-config/export")
async def submit_proxy_export(instance_id: str, body: dict, user: dict = Depends(get_current_user)):
    """提交对话导出（对接外部 EXPORT_BASE /export/submit）。

    body: {role, export_name, model?, mode?}。role 定位 create_params 里的完整 key+model，
    后端拼 access-key 调外部服务，把返回的 export_id/session_path 落 model_export_tasks（queued）。
    """
    role = (body or {}).get("role") or ""
    if role not in _ROLE_TO_KEY:
        raise HTTPException(status_code=400, detail=f"非法角色: {role}")
    export_name = (body or {}).get("export_name") or ""
    mode = (body or {}).get("mode") or _DEFAULT_MODE
    if mode not in _EXPORT_MODES:
        raise HTTPException(status_code=400, detail=f"非法 mode {mode}（allowed: {', '.join(_EXPORT_MODES)}）")

    inst = _get_instance(instance_id)
    key, model = _role_cred_from(inst.get("create_params"), role, inst)
    if (body or {}).get("model"):
        model = body["model"]

    resp = await _call_export("/export/submit", {
        "key": key,
        "model": model,
        "export_name": export_name,
        "mode": mode,
    })
    export_id = resp.get("export_id")
    session_path = resp.get("session_path") or ""

    _upsert_export_task(
        instance_id, role, key, model,
        export_id=export_id, export_name=export_name, mode=mode,
        status="queued", session_path=session_path,
    )
    return {
        "instance_id": instance_id,
        "role": role,
        "export_id": export_id,
        "session_path": session_path,
        "status": "queued",
        "mode": mode,
        "export_name": export_name,
    }


@router.get("/{instance_id}/proxy-config/export/status")
async def proxy_export_status(instance_id: str, user: dict = Depends(get_current_user)):
    """查询对话导出状态（对接外部 EXPORT_BASE /export/status）。

    只轮询 status IN (queued,running) 的行；终态/unexported 直接读表，不打外部。
    返回全部角色的 export_status + session_path + total_sessions + error_message。
    """
    inst = _get_instance(instance_id)
    export_map = _export_status_map(instance_id)

    # 刷新进行中的行：逐个调外部 status，回填 DB
    for role, ex in export_map.items():
        if (ex.get("status") or "") not in _EXPORT_STATUS_ACTIVE:
            continue
        key, model = _role_cred_from(inst.get("create_params"), role, inst)
        if not ex.get("export_id"):
            continue
        try:
            resp = await _call_export("/export/status", {
                "export_id": ex["export_id"],
                "key": key,
            })
        except HTTPException as e:
            # 外部报错（如 export_id 已清）不把任务置 failed 之外的猜：回填 error 并标 failed
            _upsert_export_task(
                instance_id, role, key, model,
                export_id=ex.get("export_id"), export_name=ex.get("export_name") or "",
                mode=ex.get("mode") or _DEFAULT_MODE, status="failed",
                session_path=ex.get("session_path") or "",
                error_message=str(e.detail),
            )
            continue
        status = (resp or {}).get("status") or "failed"
        _upsert_export_task(
            instance_id, role, key, model,
            export_id=ex.get("export_id"), export_name=ex.get("export_name") or "",
            mode=ex.get("mode") or _DEFAULT_MODE, status=status,
            session_path=(resp or {}).get("session_path") or ex.get("session_path") or "",
            total_sessions=(resp or {}).get("total_sessions"),
            error_message=(resp or {}).get("error_message") or "",
        )

    # 重新读最新状态返回
    export_map = _export_status_map(instance_id)
    rows = []
    cp = None
    try:
        if inst.get("create_params"):
            cp = json.loads(inst["create_params"])
    except (ValueError, TypeError):
        cp = None
    for r in _build_proxy_rows(cp):
        ex = export_map.get(r["role"])
        rows.append({
            "role": r["role"],
            "export_status": (ex or {}).get("status") or "unexported",
            "session_path": (ex or {}).get("session_path") or "",
            "total_sessions": (ex or {}).get("total_sessions"),
            "export_id": (ex or {}).get("export_id"),
            "error_message": (ex or {}).get("error_message") or "",
        })
    return {"instance_id": instance_id, "rows": rows}


@router.get("/{instance_id}/proxy-config/browse")
async def browse_proxy_export(instance_id: str, role: str = Query(...),
                              user: dict = Depends(get_current_user)):
    """返回外部对话浏览页 URL（前端拿到后 window.open 到 export_base/export/view）。

    完整 key + access-key 拼在后端 URL 里，前端只拿成品链接不可见 key 明文。
    不能直接用 302 重定向：本端点是鉴权接口，window.open 的裸导航不带 Bearer token 会 401，
    永远到不了重定向。改为返回 JSON，前端走带 token 的 axios 取 URL 再新开标签页。
    """
    inst = _get_instance(instance_id)
    key, model = _role_cred_from(inst.get("create_params"), role, inst)
    # 外部 /export/view 需要 access-key 参数；_call_export 自动附加，但 browse 走 _export_url
    # 直接拼链接，这里显式带上 access-key。
    url = _export_url("/export/view", {
        "key": key,
        "model": model,
        "access-key": settings.EXPORT_ACCESS_KEY,
    })
    return {"url": url}
